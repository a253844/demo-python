from pickle import FALSE, TRUE
import openpyxl 
import os 
import customtkinter as ctk
from tkinter import filedialog, messagebox

hasNumber = lambda stringVal: any(elem.isdigit() for elem in stringVal)

def get_sheet_names(filepath):
    try:
        # Load workbook in read_only mode to simply fetch sheet names efficiently
        wb = openpyxl.load_workbook(filepath, read_only=True)
        return wb.sheetnames
    except Exception as e:
        return []

def TranslateData(file_path, sheet_name, year, month, log_callback):
    try:
        if not os.path.isfile("Sample.xlsx"):
            log_callback("哭哭，寶貝找不到範本檔(Sample.xlsx) ，幫我檢查一下 ಥ⌣ಥ\n")
            return False

        log_callback(f"開始讀取 {os.path.basename(file_path)}...\n")
        Datawb = openpyxl.load_workbook(file_path)
        Samplewb = openpyxl.load_workbook('Sample.xlsx')

        if sheet_name not in Datawb.sheetnames:
            log_callback(f"{sheet_name} 資料表，寶貝找不到 (@~@)?，我要哭了喔\n")
            return False
            
        Datasheet = Datawb[sheet_name]
        Samplesheet = Samplewb['工作表1']

        # 取出欄位
        DataTitleName = Datasheet['1']
        SampleTitleName = Samplesheet['1']
        UserName = any
        UserId = any
        UserIdType = any
        DateRange = any
        DateStart = ''
        DateEnd = ''

        # 找出須處理的資料列位置 
        for cell in DataTitleName:
            if cell.value == '身分證':
                UserId = Datasheet[cell.column_letter]
            if cell.value == '身分別':
                UserIdType = Datasheet[cell.column_letter]
            if cell.value == '姓名日期':
                UserName = Datasheet[cell.column_letter]
            if cell.value is not None:
                if type(cell.value) == int:
                    if DateStart == '':
                        DateStart = cell.column_letter
                    else:
                        DateEnd = cell.column_letter
                    
        if DateStart == '' or DateEnd == '':
            log_callback("警告：在第一列找不到代表日期的數字欄位。\n")
            return False

        DateRange = Datasheet[DateStart + ":" + DateEnd]
        DataList = []

        log_callback("開始過濾與提取資料...\n")
        for cell in UserId:
            if cell.value is not None:
                if hasNumber(cell.value): 
                    rownum = cell.row
                    TypeName = UserIdType[rownum -1].value
                    TempUserName = UserName[rownum -1].value
                    UType = 0
                    tempList = []
                    
                    if TypeName is not None:
                        if TypeName.find("低收") != -1:
                            UType = 1
                        if TypeName.find("一般戶") != -1:
                            UType = 2

                    if UType != 0:
                        for i in DateRange:
                            row_idx = rownum
                            col_idx = i[0].column
                            Data = Datasheet.cell(row=row_idx, column=col_idx)
                            if Data.value == 1:
                                datenum = str(i[0].value) if len(str(i[0].value)) > 1 else "0" + str(i[0].value)
                                List = []
                                List.append(cell.value)
                                List.append(str(year) + str(month).zfill(2) + datenum)
                                List.append(UType)
                                List.append(TempUserName)
                                tempList.append(List)
                    if tempList:
                        DataList.append(tempList)

        count = 2
        for j in DataList:
            for k in j:
                # 身分證字號
                Samplesheet.cell(row=count, column=1, value=k[0])
                # 服務日期
                Samplesheet.cell(row=count, column=2, value=k[1])
                # 服務項目代碼
                Samplesheet.cell(row=count, column=3, value="OT01")
                # 服務類別
                Samplesheet.cell(row=count, column=4, value=k[2])
                # 备注
                Samplesheet.cell(row=count, column=12, value=k[3])
                # 數量
                Samplesheet.cell(row=count, column=5, value="1")
                # 單價
                if k[2] == 1:
                    Samplesheet.cell(row=count, column=6, value="100")
                else:
                    Samplesheet.cell(row=count, column=6, value="90")
                
                # 服務人員身分證
                Samplesheet.cell(row=count, column=7, value="A229516580")
                # 起始時段-小時
                Samplesheet.cell(row=count, column=8, value="11")
                # 起始時段-分鐘
                Samplesheet.cell(row=count, column=9, value="0")
                # 結束時段-小時
                Samplesheet.cell(row=count, column=10, value="11")
                # 結束時段-分鐘
                Samplesheet.cell(row=count, column=11, value="30")
                # OT01必填-餐別
                Samplesheet.cell(row=count, column=24, value="2")
                count += 1
                
        Samplewb.save("CompleteData.xlsx")
        
        # 將 xlsx 轉換為 97-2003 xls 格式
        try:
            import win32com.client
            excel = win32com.client.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # 轉換需要使用絕對路徑
            xlsx_path = os.path.abspath("CompleteData.xlsx")
            xls_path = os.path.abspath("CompleteData.xls")
            
            if os.path.exists(xls_path):
                try:
                    os.remove(xls_path)
                except Exception:
                    pass
                
            wb = excel.Workbooks.Open(xlsx_path)
            # 56 代表 xlExcel8 (97-2003 format in Excel)
            wb.SaveAs(xls_path, FileFormat=56)
            wb.Close()
            excel.Quit()
            
            # 若轉換成功，可以選擇刪除暫存的 xlsx
            try:
                os.remove(xlsx_path)
            except Exception:
                pass
                
            log_callback("寶貝幫你把資料成功轉換好了，檔案已另存囉，記住檔案名稱叫 CompleteData.xls\n")
            log_callback("我是不是很棒，給我一百個抱抱!!!!  (~￣▽￣)~\n")
        except Exception as e:
            # 發生例外 (例如尚未安裝 pywin32，或使用者電腦沒安裝 Excel)
            log_callback(f"⚠️ 轉換 97-2003 格式 (.xls) 時發生錯誤：{str(e)}\n")
            log_callback("寶貝幫你把資料成功轉換好了，請使用備用檔案 CompleteData.xlsx\n")
            log_callback("我是不是很棒，給我一百個抱抱!!!!  (~￣▽￣)~\n")
            
        return True
    except Exception as e:
        log_callback(f"發生錯誤了 ಥ⌣ಥ : {str(e)}\n")
        return False

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("Excel Data Converter 小幫手")
        self.geometry("600x550")
        self.grid_columnconfigure(0, weight=1)
        
        # 標題
        self.title_label = ctk.CTkLabel(self, text="🌟 試算表轉換小工具 🌟", font=ctk.CTkFont(size=24, weight="bold"))
        self.title_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        
        # 主框架
        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        self.main_frame.grid_columnconfigure(1, weight=1)
        
        # 檔案選取
        self.file_label = ctk.CTkLabel(self.main_frame, text="1. 選擇檔案:")
        self.file_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        self.file_path_var = ctk.StringVar()
        self.file_entry = ctk.CTkEntry(self.main_frame, textvariable=self.file_path_var, state="readonly")
        self.file_entry.grid(row=0, column=1, padx=(0, 10), pady=10, sticky="ew")
        
        self.browse_btn = ctk.CTkButton(self.main_frame, text="瀏覽...", command=self.browse_file, width=80)
        self.browse_btn.grid(row=0, column=2, padx=(0, 10), pady=10)
        
        # 工作表選取
        self.sheet_label = ctk.CTkLabel(self.main_frame, text="2. 工作表:")
        self.sheet_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        
        self.sheet_var = ctk.StringVar()
        self.sheet_dropdown = ctk.CTkComboBox(self.main_frame, variable=self.sheet_var, values=["請先選擇檔案"], state="readonly")
        self.sheet_dropdown.grid(row=1, column=1, columnspan=2, padx=(0, 10), pady=10, sticky="ew")
        
        # 年份與月份
        self.year_label = ctk.CTkLabel(self.main_frame, text="3. 年份 (民國):")
        self.year_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")
        
        self.year_entry = ctk.CTkEntry(self.main_frame, placeholder_text="例如: 111")
        self.year_entry.grid(row=2, column=1, columnspan=2, padx=(0, 10), pady=10, sticky="ew")
        
        self.month_label = ctk.CTkLabel(self.main_frame, text="4. 月份:")
        self.month_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")
        
        self.month_entry = ctk.CTkEntry(self.main_frame, placeholder_text="例如: 5 或 05")
        self.month_entry.grid(row=3, column=1, columnspan=2, padx=(0, 10), pady=10, sticky="ew")

        # 執行區塊
        self.execute_btn = ctk.CTkButton(self, text="🚀 開始轉換資料", command=self.start_processing, font=ctk.CTkFont(size=16, weight="bold"), height=40)
        self.execute_btn.grid(row=2, column=0, padx=20, pady=15, sticky="ew")
        
        # 日誌輸出
        self.log_textbox = ctk.CTkTextbox(self, height=150)
        self.log_textbox.grid(row=3, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.log_textbox.insert("0.0", "歡迎使用轉換小幫手！請按步驟選擇檔案並填寫資料。\n")
        self.log_textbox.configure(state="disabled")

    def log_message(self, message):
        self.log_textbox.configure(state="normal")
        self.log_textbox.insert("end", message)
        self.log_textbox.see("end")
        self.log_textbox.configure(state="disabled")
        self.update()

    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="選擇 Excel 檔案",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if filename:
            self.file_path_var.set(filename)
            self.log_message(f"已選擇檔案: {os.path.basename(filename)}\n讀取工作表中...\n")
            
            sheets = get_sheet_names(filename)
            if sheets:
                self.sheet_dropdown.configure(values=sheets)
                self.sheet_var.set(sheets[0])
                self.log_message("讀取工作表完成。請輸入年份與月份。\n\n")
            else:
                self.sheet_dropdown.configure(values=["讀取失敗"])
                self.sheet_var.set("讀取失敗")
                self.log_message("讀取工作表失敗，請確定這是一個正確的 Excel 檔案。\n\n")

    def start_processing(self):
        file_path = self.file_path_var.get()
        sheet_name = self.sheet_var.get()
        year = self.year_entry.get().strip()
        month = self.month_entry.get().strip()
        
        if not file_path:
            self.log_message("❌ 請先選擇檔案！\n")
            return
            
        if sheet_name in ["", "請先選擇檔案", "讀取失敗"]:
            self.log_message("❌ 請選擇正確的工作表！\n")
            return
            
        if not year or not month:
            self.log_message("❌ 請輸入年份與月份！\n")
            return
            
        self.log_message("-" * 40 + "\n")
        self.log_message(f"🚀 開始執行任務 (工作表: {sheet_name})...\n")
        self.execute_btn.configure(state="disabled")
        self.update()
        
        success = TranslateData(file_path, sheet_name, year, month, self.log_message)
        
        if success:
            self.log_message("✅ 任務圓滿完成！\n\n")
        else:
            self.log_message("❌ 任務失敗或是中止。\n\n")
            
        self.execute_btn.configure(state="normal")

if __name__ == "__main__":
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")
    
    app = App()
    app.mainloop()
